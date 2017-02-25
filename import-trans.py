import sys
import re
import csv
import datetime
import openpyxl
import argparse
from constants import *

# Process input arguments
parser = argparse.ArgumentParser()
parser.add_argument("tfile", help="transaction file name")
parser.add_argument("rfile", help="raw file name to import into transaction file")
args = parser.parse_args()


# Read in data from transaction file
wb_trans = openpyxl.load_workbook(args.tfile)

journal_lod = list()
sheet = wb_trans.get_sheet_by_name('Journal')

# Check header row to make sure it is correct
#row = sheet[1]
#print(row)
#for cell in row:
#    print(cell.value)

#for rowNum in range(2, sheet.max_row + 1):  # skip the first row, since it is h
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):  # Skip first row
    journal_lod.append({journal_header_list[0]: row[0].value,
                        journal_header_list[1]: row[1].value,
                        journal_header_list[2]: row[2].value,
                        journal_header_list[3]: row[3].value,
                        journal_header_list[4]: row[4].value,
                        journal_header_list[5]: row[5].value,
                        journal_header_list[6]: row[6].value,
                        journal_header_list[7]: row[7].value,
                        journal_header_list[8]: row[8].value,
                        journal_header_list[9]: row[9].value,
                        journal_header_list[10]: row[10].value,
                        journal_header_list[11]: row[11].value,
                        journal_header_list[12]: row[12].value,
                        journal_header_list[13]: row[13].value,
                        journal_header_list[14]: row[14].value})


raw_lod = list()
sheet = wb_trans.get_sheet_by_name('Raw')

# Check header row to make sure it is correct
#row = sheet[1]
#print(row)
#for cell in row:
#    print(cell.value)

# Skip header row (first row)
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
    raw_lod.append({raw_header_list[0]: row[0].value,
                    raw_header_list[1]: row[1].value,
                    raw_header_list[2]: row[2].value,
                    raw_header_list[3]: row[3].value})

# Sort transaction raw data by RefNo
raw_lod.sort(key=lambda r: r[raw_header_list[0]])

# Extract reference number data
mobj = re.search(r"(\w+-\d+)-(\d+)", raw_lod[-1][raw_header_list[0]])
if mobj:
    control_refno_format = mobj.group(1)
    control_refno_max = int(mobj.group(2))
else:
    sys.exit("Invalid reference number format '{}'".format(raw_lod[-1][raw_header_list[0]]))


control_kw_lod = list()
sheet = wb_trans.get_sheet_by_name('Control')

control_file_dl_instr = sheet["B1"].value
control_data_start_line = sheet["B5"].value
control_col_date = sheet["B6"].value
control_col_amnt = sheet["B7"].value
control_col_desc = sheet["B8"].value
control_date_format = sheet["B9"].value

# Start at first data row of keyword table
for row in sheet.iter_rows(min_row=14, max_row=sheet.max_row):
    control_kw_lod.append({keyword_header_list[0]: row[0].value,
                           keyword_header_list[1]: row[1].value,
                           keyword_header_list[2]: row[2].value,
                           keyword_header_list[3]: row[3].value,
                           keyword_header_list[4]: row[4].value,
                           keyword_header_list[5]: row[5].value,
                           keyword_header_list[6]: row[6].value,
                           keyword_header_list[7]: row[7].value,
                           keyword_header_list[8]: row[8].value})



# Read in data from raw import file
import_raw_lod = list()
inf = open(args.rfile, newline='')
reader = csv.reader(inf)

row_cnt = 0
for row in reader:
    row_cnt += 1

    if row_cnt >= control_data_start_line:
        date = datetime.datetime.strptime(row[control_col_date - 1], control_date_format)
        amnt = float(row[control_col_amnt - 1])
        desc = row[control_col_desc - 1]

        import_raw_lod.append({raw_header_list[1]: date,
                               raw_header_list[2]: amnt,
                               raw_header_list[3]: desc})

inf.close()

# Check for duplicate data in imported raw data
for i in range(0, len(import_raw_lod)):
    for j in range(0, len(import_raw_lod)):
        if i != j:
            if ((import_raw_lod[i][raw_header_list[1]] == import_raw_lod[j][raw_header_list[1]]) and
                (import_raw_lod[i][raw_header_list[2]] == import_raw_lod[j][raw_header_list[2]]) and
                (import_raw_lod[i][raw_header_list[3]] == import_raw_lod[j][raw_header_list[3]])):
                print_string = "Found duplicate transaction in import file '{}' at rows {} and {}"
                sys.exit(print_string.format(args.rfile, control_data_start_line + i, control_data_start_line + j))

# Sort imported raw data by Date
import_raw_lod.sort(key=lambda r: r[raw_header_list[1]])

# Compare rows from imported raw transactions with those already the raw transaction list, if new, add to raw
# transaction list
add_raw_lod = list()
for import_row in import_raw_lod:
    already_in_table = False
    for trans_row in raw_lod:
        if ((import_row[raw_header_list[1]] == trans_row[raw_header_list[1]]) and
            (import_row[raw_header_list[2]] == trans_row[raw_header_list[2]]) and
            (import_row[raw_header_list[3]] == trans_row[raw_header_list[3]])):
            already_in_table = True
            break

    if not already_in_table:
        control_refno_max += 1
        import_row[raw_header_list[0]] = "{}-{}".format(control_refno_format, control_refno_max)
        raw_lod.append(import_row)
        add_raw_lod.append(import_row)

# Sort imported journal transaction data by RefNo
journal_lod.sort(key=lambda r: r[journal_header_list[1]])

# Convert additional raw lod transactions to additional journal lod transactions by seeing if descriptions match those
# in key word table
add_journal_lod = list()
for raw_row in add_raw_lod:
    date = raw_row[raw_header_list[1]]
    rfno = raw_row[raw_header_list[0]]
    amnt = raw_row[raw_header_list[2]]
    desc = raw_row[raw_header_list[3]]
    dupl = ""
    srce = ""
    spln = ""
    dr_mrch = ""
    dr_atyp = ""
    dr_acnt = ""
    dr_suba = ""
    cr_mrch = ""
    cr_atyp = ""
    cr_acnt = ""
    cr_suba = ""

    for key_row in control_kw_lod:
        mobj = re.search(key_row[keyword_header_list[0]], desc)
        if mobj:
            dr_mrch = key_row[keyword_header_list[1]]
            dr_atyp = key_row[keyword_header_list[2]]
            dr_acnt = key_row[keyword_header_list[3]]
            dr_suba = key_row[keyword_header_list[4]]
            cr_mrch = key_row[keyword_header_list[5]]
            cr_atyp = key_row[keyword_header_list[6]]
            cr_acnt = key_row[keyword_header_list[7]]
            cr_suba = key_row[keyword_header_list[8]]
            break

    journal_row = {journal_header_list[0]: date,
                   journal_header_list[1]: rfno,
                   journal_header_list[2]: dr_mrch,
                   journal_header_list[3]: dr_atyp,
                   journal_header_list[4]: dr_acnt,
                   journal_header_list[5]: dr_suba,
                   journal_header_list[6]: amnt,
                   journal_header_list[7]: cr_mrch,
                   journal_header_list[8]: cr_atyp,
                   journal_header_list[9]: cr_acnt,
                   journal_header_list[10]: cr_suba,
                   journal_header_list[11]: desc,
                   journal_header_list[12]: dupl,
                   journal_header_list[13]: srce,
                   journal_header_list[14]: spln}

    add_journal_lod.append(journal_row)

# Add additional lod to journal lod
journal_lod.extend(add_journal_lod)


# Write additional raw lod to excel file
sheet = wb_trans.get_sheet_by_name('Raw')
sheet.freeze_panes = 'A2'
row_cnt = sheet.max_row + 1
for row in add_raw_lod:
    col_cnt = 1
    for item in raw_header_list:
        sheet.cell(row=row_cnt, column=col_cnt).value = row[item]
        col_cnt += 1

    row_cnt += 1

# Write additional lod journal to excel file
sheet = wb_trans.get_sheet_by_name('Journal')
sheet.freeze_panes = 'A2'
row_cnt = sheet.max_row + 1
for row in add_journal_lod:
    col_cnt = 1
    for item in journal_header_list:
        sheet.cell(row=row_cnt, column=col_cnt).value = row[item]
        col_cnt += 1

    row_cnt += 1

# Write to excel file to same file name
wb_trans.save(args.tfile)

# Temporary code to write csv file for output
file_handle = open("test_out.csv", 'w', newline='')
writer = csv.DictWriter(file_handle, journal_header_list, quoting=csv.QUOTE_MINIMAL)
writer.writeheader()
writer.writerows(add_journal_lod)
file_handle.close()

print("Done")