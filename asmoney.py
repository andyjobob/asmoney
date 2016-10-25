import sys
import os
import re
import openpyxl
import transaction

################################################################################
# Description:
#
#  Arguments:
#    1. dir      - Directory where transactions are stored. Make sure dir only
#                  contains transactions (as *.xlsx) and no sub directories
#    2. out_file - File name to write output data to
#
# ex. > python asmoney.py ./dir out_file.csv
#
# Revision:
#   0.1, Andy Sciullo, 09/01/2016, Initial Version
#
################################################################################
################################################################################
# Notes:
#   - Files must be named "*_Transactions.xlsx"
#   - Workbooks should have all transactions in a sheet named "Journal"
#
################################################################################

################################################################################
### Constants
################################################################################
COL_DATE = 1
COL_RFNO = 2
COL_DRMRCH = 3
COL_DRATYP = 4
COL_DRACNT = 5
COL_DRASUB = 6
COL_AMNT = 7
COL_CRMRCH = 8
COL_CRATYP = 9
COL_CRACNT = 10
COL_CRASUB = 11
COL_DESC = 12
COL_DUPL = 13
COL_DSRC = 14
COL_SPLIT = 15

################################################################################
### Global Variables
################################################################################
dictDrDate = dict()
dictDrMrch = dict()
dictDrAtyp = dict()
dictDrAcnt = dict()
dictDrAsub = dict()
dictDrAmnt = dict()
dictDrDesc = dict()
dictCrDate = dict()
dictCrMrch = dict()
dictCrAtyp = dict()
dictCrAcnt = dict()
dictCrAsub = dict()
dictCrAmnt = dict()
dictCrDesc = dict()
arrDrRefNo = []
arrCrRefNo = []
trans_dict = dict()

################################################################################
# Function process_workbook
#
# Description:
#   Check the given file name for correct naming format and read contents for
#   BFM data.
#
#  Arguments:
#    1. file_name - Name of file to process.
#
#  Return Values:
#    None
#
################################################################################
def process_workbook(wb_name):
    # Init variables
    trans_list = list()

    # Search for files with correct name format
    print("  Processing Workbook: ", wb_name)
    wb_temp = openpyxl.load_workbook(os.path.join(dir, wb_name))
    sheet = wb_temp.get_sheet_by_name('Journal')
    for rowNum in range(2, sheet.max_row + 1):  # skip the first row
        temp_dict = None
        temp_dict = dict()
        refno = sheet.cell(row=rowNum, column=COL_RFNO).value
        amnt = sheet.cell(row=rowNum, column=COL_AMNT).value
        dupl = sheet.cell(row=rowNum, column=COL_DUPL).value
        tsplit = sheet.cell(row=rowNum, column=COL_SPLIT).value
        refsplit = refno + "-" + tsplit
        if not dupl:
            if refsplit in trans_dict.keys():
                print("    Duplicate reference/split number found -  RFNO: {0}, SPLIT: {1}, AMNT: {2}".format(refno, tsplit, amnt))
            else:
                temp_dict["date"] = sheet.cell(row=rowNum, column=COL_DATE).value
                temp_dict["refno"] = refno
                temp_dict["drmrch"] = sheet.cell(row=rowNum, column=COL_DRMRCH).value
                temp_dict["dratyp"] = sheet.cell(row=rowNum, column=COL_DRATYP).value
                temp_dict["dractn"] = sheet.cell(row=rowNum, column=COL_DRACNT).value
                temp_dict["drasub"] = sheet.cell(row=rowNum, column=COL_DRASUB).value
                temp_dict["amnt"] = amnt
                temp_dict["crmrch"] = sheet.cell(row=rowNum, column=COL_CRMRCH).value
                temp_dict["cratyp"] = sheet.cell(row=rowNum, column=COL_CRATYP).value
                temp_dict["cractn"] = sheet.cell(row=rowNum, column=COL_CRACNT).value
                temp_dict["crasub"] = sheet.cell(row=rowNum, column=COL_CRASUB).value
                temp_dict["desc"] = sheet.cell(row=rowNum, column=COL_DESC).value
                temp_dict["dupl"] = dupl
                temp_dict["date"] = sheet.cell(row=rowNum, column=COL_DSRC).value
                temp_dict["tsplit"] = tsplit
                trans_dict[refsplit] = temp_dict
        #else:
        #    print "    Excluding marked duplicate - RFNO: ", refNo, ", AMNT: ", amnt

################################################################################
# Function print_transactions
#
# Description:
#   Check the given file name for correct naming format and read contents for
#   BFM data.
#
#  Arguments:
#    1. file_name - Name of file to process.
#
#  Return Values:
#    None
#
################################################################################
def print_transactions(wb_name):
    print("\n  Writing Output: {0}".format(wb_name))

    wb_out = openpyxl.load_workbook(wb_name)
    sheet = wb_out.get_sheet_by_name("Journal")

    # Clear out previous journal data
    for rowNum in range(2, sheet.max_row + 1):  # skip the first row
        for colNum in range(1, sheet.max_column + 1):
            sheet.cell(row=rowNum, column=colNum).value = ""

    rowCnt = 2
    for ref in trans_dict.keys():
        #print("%s,%s,%s,%s,%s,%s,%s,%s" % (refNo, dictDrDate[refNo], dictDrMrch[refNo], dictDrAtyp[refNo], dictDrAcnt[refNo], dictDrAsub[refNo], \
        #    dictDrAmnt[refNo], dictDrDesc[refNo]))
        #print("%s,%s,%s,%s,%s,%s,%s,%s" % (refNo, dictCrDate[refNo], dictCrMrch[refNo], dictCrAtyp[refNo], dictCrAcnt[refNo], dictCrAsub[refNo], \
        #    dictCrAmnt[refNo], dictCrDesc[refNo]))

        # Debit row
        sheet.cell(row=rowCnt, column=COL_DATE).value = trans_dict[ref]["date"]
        sheet.cell(row=rowCnt, column=COL_RFNO).value = trans_dict[ref]["refno"]
        sheet.cell(row=rowCnt, column=COL_MRCH).value = trans_dict[ref]["drmrch"]
        sheet.cell(row=rowCnt, column=COL_ATYP).value = trans_dict[ref]["dratyp"]
        sheet.cell(row=rowCnt, column=COL_ACNT).value = trans_dict[ref]["dracnt"]
        sheet.cell(row=rowCnt, column=COL_ASUB).value = trans_dict[ref]["drasub"]
        sheet.cell(row=rowCnt, column=COL_AMNT).value = trans_dict[ref]["amnt"]
        sheet.cell(row=rowCnt, column=COL_DESC).value = trans_dict[ref]["desc"]
        rowCnt += 1

        # Credit row
        sheet.cell(row=rowCnt, column=COL_DATE).value = trans_dict[ref]["date"]
        sheet.cell(row=rowCnt, column=COL_RFNO).value = trans_dict[ref]["refno"]
        sheet.cell(row=rowCnt, column=COL_MRCH).value = trans_dict[ref]["crmrch"]
        sheet.cell(row=rowCnt, column=COL_ATYP).value = trans_dict[ref]["cratyp"]
        sheet.cell(row=rowCnt, column=COL_ACNT).value = trans_dict[ref]["cracnt"]
        sheet.cell(row=rowCnt, column=COL_ASUB).value = trans_dict[ref]["crasub"]
        sheet.cell(row=rowCnt, column=COL_AMNT).value = trans_dict[ref]["amnt"] * -1
        sheet.cell(row=rowCnt, column=COL_DESC).value = trans_dict[ref]["desc"]
        rowCnt += 1

    sheet.freeze_panes = 'A2'
    wb_out.save(wb_name)


################################################################################
# Main
#
# Description:
#   Check the given file name for correct naming format and read contents for
#   BFM data.
#
#  Arguments:
#    1. file_name - Name of file to process.
#
#  Return Values:
#    None
#
################################################################################
# Process input arguments
if (len(sys.argv) != 3):
    print("Usage: python asmoney.py <PATH_TO_DIR> <OUTFILE>")
    sys.exit()

dir = str(sys.argv[1])
out_file = str(sys.argv[2])

print("Transaction Dir : {0}".format(dir))
print("Output File     : {0}".format(out_file))

# Loop through each file in directory
for file_name in os.listdir(dir):
    if os.path.isfile(os.path.join(dir, file_name)):
        search_obj = re.search(r".*_Transactions\.xlsx", file_name)
        if search_obj:
            process_workbook(file_name)

print_transactions(out_file)
print("\n ==> ...Ending")








